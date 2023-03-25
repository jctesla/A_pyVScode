# This example copies the values from the source range 'A1:C3' to the destination range 'E5:G7' and then clear the source range.
import openpyxl
workbook = openpyxl.load_workbook('col1_col2.xlsx',read_only=False)           # Load the Excel workbook
print(f"1. This is the sheets for this excel = {workbook.sheetnames}")
wb_sht = workbook.sheetnames
#print(dir(workbook))

sheet = workbook[wb_sht[0]]                                  # Select sheet[0] = 'words_ingles'
#sheet = workbook.active                                     # Select the active sheet

src_cell = sheet['A2:A151']                                  # Define the source cell (the cell to be moved)
des_cell = sheet['C2:C151']                                  # Define the destination cell (where the source cell will be moved to) 

print(type(src_cell),"-",type(src_cell))
print(len(src_cell),"-",len(des_cell) )
# print(src_cell[:5])


# Iterate over the source range and copy the values to the destination range
# the structure if each tupla, are = (cell(0),cell(1),cell(2)....) :. cell(0)(0),cell(0)(1),cell(0)(2) ...
for source_row, destination_row in zip(src_cell, des_cell):
    for source_cell, destination_cell in zip(source_row, destination_row):
        destination_cell.value = source_cell.value
        source_cell.value = None                             # borro la celda origen       

workbook.save('col1_col2.xlsx')                                # Save the changes to the Excel file 

"""
# Clear the source range
for row in source_range:
    for cell in row:
        cell.value = None
"""